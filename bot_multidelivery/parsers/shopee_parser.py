"""
Parser Excel da Shopee - Extrai entregas do romaneio oficial
"""
from typing import List, Dict
from dataclasses import dataclass
import openpyxl
import re


def clean_destination_address(raw_address: str) -> str:
    """
    Limpa endereço da Shopee extraindo APENAS:
    - Nome da rua (antes da primeira vírgula)
    - Número do prédio (após a primeira vírgula, até encontrar espaço/vírgula/parêntese)
    
    Exemplos:
        "Rua Mena Barreto, 151, Portaria" -> "Rua Mena Barreto, 151"
        "Rua Principado de Mônaco, 37, Apt 501(guarita tb pode deixar" -> "Rua Principado de Mônaco, 37"
        "Rua Real Grandeza, 278, 601" -> "Rua Real Grandeza, 278"
    """
    if not raw_address:
        return ""
    
    # Remove espaços extras
    address = raw_address.strip()
    
    # Divide pela primeira vírgula
    parts = address.split(',', 2)  # Limita a 3 partes
    
    if len(parts) < 2:
        # Se não tem vírgula, retorna o endereço como está
        return address
    
    # Parte 1: Nome da rua
    street_name = parts[0].strip()
    
    # [FIX] Remove lixo comum que aparece no nome da rua (antes da vírgula)
    # Ex: "Rua X apt 201" -> "Rua X"
    # Remove qualquer coisa que pareça "apt", "ap", "bloco", "loja" seguido de digitos ou no final
    street_name = re.sub(r'\s+(?:apt\.?|ap\.?|apto\.?|bloco|bl\.?|loja|lj\.?|casa|sl\.?|sala)\s*.*$', '', street_name, flags=re.IGNORECASE)

    # Parte 2: Número do prédio (extrai apenas dígitos do início)
    number_part = parts[1].strip()
    
    # Extrai apenas o número (remove tudo após espaços, parênteses, vírgulas)
    number_match = re.match(r'^(\d+[A-Za-z]?)', number_part)
    if number_match:
        building_number = number_match.group(1)
    else:
        # Se não encontrar número, usa a parte toda
        building_number = number_part.split()[0] if ' ' in number_part else number_part
    
    # Retorna apenas rua + número
    return f"{street_name}, {building_number}"


@dataclass
class ShopeeDelivery:
    """Entrega individual do romaneio Shopee"""
    tracking: str
    address: str
    bairro: str
    city: str
    latitude: float
    longitude: float
    stop: int
    customer_name: str = ""
    phone: str = ""


def parse_shopee_excel(file_path: str) -> List[Dict[str, any]]:
    """
    Parse Excel da Shopee (formato DD-MM-YYYY Nome.xlsx)
    
    Extrai:
    - Tracking code
    - Endereço completo
    - Lat/Lon (embutidos na planilha)
    - STOP (agrupamento de mesmo prédio)
    
    Returns:
        Lista de dicts com: id, address, lat, lon, stop
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Procura cabeçalhos (geralmente na linha 1 ou 2)
        headers = {}
        header_row = None
        
        for row in range(1, min(5, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    
                    # Tracking: SPX TN, tracking, código, AT ID
                    if any(x in cell_lower for x in ['spx tn', 'spx_tn', 'tracking', 'código', 'at id', 'atid']):
                        headers['tracking'] = col
                        header_row = row
                    # Endereço: Destination, address, endereço
                    elif any(x in cell_lower for x in ['destination', 'endereço', 'endereco', 'address']):
                        headers['address'] = col
                        if header_row is None:  # Define header_row se ainda não definido
                            header_row = row
                    # Rua: planilhas simplificadas
                    elif any(x in cell_lower for x in ['rua', 'street']):
                        headers['street'] = col
                        if header_row is None:
                            header_row = row
                    # Bairro: neighborhood, bairro, distrito, district
                    elif any(x in cell_lower for x in ['bairro', 'distrito', 'district', 'neighborhood', 'neighbour']):
                        headers['bairro'] = col
                        if header_row is None:
                            header_row = row
                    # Cidade
                    elif 'city' in cell_lower or 'cidade' in cell_lower:
                        headers['city'] = col
                    # Latitude
                    elif 'latitude' in cell_lower or (cell_lower == 'lat'):
                        headers['lat'] = col
                    # Longitude
                    elif 'longitude' in cell_lower or (cell_lower in ['lng', 'lon', 'long']):
                        headers['lon'] = col
                    # Stop (Sequence ou Stop)
                    elif any(x in cell_lower for x in ['stop', 'parada', 'sequence']):
                        headers['stop'] = col
                    # Cliente
                    elif any(x in cell_lower for x in ['nome', 'cliente', 'customer', 'name']):
                        headers['customer'] = col
                    # Telefone
                    elif any(x in cell_lower for x in ['telefone', 'phone', 'tel']):
                        headers['phone'] = col
        
        # Validação flexível: precisa de pelo menos tracking OU address
        if not header_row:
            raise ValueError("Cabeçalhos não encontrados. Formato inválido.")
        
        if 'tracking' not in headers and 'address' not in headers and 'street' not in headers:
            raise ValueError("Não encontrei coluna de Tracking (SPX TN), Endereço (Destination) ou Rua.")
        
        # Extrai dados
        addresses = []
        stop_counter = 1
        
        for row in range(header_row + 1, ws.max_row + 1):
            # Pega tracking (pode ser SPX TN ou outra coluna)
            tracking = None
            if 'tracking' in headers:
                tracking_cell = ws.cell(row, headers['tracking'])
                if tracking_cell.value:
                    tracking = str(tracking_cell.value).strip()
            
            # Pega endereço (Destination) ou monta a partir de rua + bairro
            address = ""
            if 'address' in headers:
                addr_cell = ws.cell(row, headers['address'])
                if addr_cell.value:
                    address = str(addr_cell.value).strip()
            if not address and 'street' in headers:
                street_cell = ws.cell(row, headers['street'])
                street_val = str(street_cell.value or '').strip()
                bairro_val = ""
                if 'bairro' in headers:
                    bairro_cell = ws.cell(row, headers['bairro'])
                    bairro_val = str(bairro_cell.value or '').strip()
                if street_val:
                    address = street_val if not bairro_val else f"{street_val}, {bairro_val}"
            
            # Se não tem tracking nem endereço, pula linha
            if not tracking and not address:
                continue
            
            # Se não tem tracking, gera um ID
            if not tracking:
                tracking = f"PKG{row:04d}"
            
            bairro = ""
            if 'bairro' in headers:
                bairro_cell = ws.cell(row, headers['bairro'])
                if bairro_cell.value:
                    bairro = str(bairro_cell.value).strip()
            
            city = ""
            if 'city' in headers:
                city_cell = ws.cell(row, headers['city'])
                if city_cell.value:
                    city = str(city_cell.value).strip()
            
            # Lat/Lon (podem estar embutidos ou precisar geocoding)
            lat = None
            lon = None
            
            if 'lat' in headers:
                lat_cell = ws.cell(row, headers['lat']).value
                if lat_cell:
                    try:
                        lat = float(lat_cell)
                    except:
                        pass
            
            if 'lon' in headers:
                lon_cell = ws.cell(row, headers['lon']).value
                if lon_cell:
                    try:
                        lon = float(lon_cell)
                    except:
                        pass
            
            # STOP (agrupamento)
            stop = stop_counter
            if 'stop' in headers:
                stop_cell = ws.cell(row, headers['stop']).value
                if stop_cell:
                    try:
                        stop = int(stop_cell)
                    except:
                        pass
            
            # Cliente e telefone (opcional)
            customer = ""
            if 'customer' in headers:
                customer = str(ws.cell(row, headers['customer']).value or '').strip()
            
            phone = ""
            if 'phone' in headers:
                phone = str(ws.cell(row, headers['phone']).value or '').strip()
            
            if address:  # Só adiciona se tem endereço
                # Limpa o endereço para geocoding (apenas rua + número)
                cleaned_address = clean_destination_address(address)
                
                addresses.append({
                    'id': tracking,
                    'address': cleaned_address,  # Endereço limpo
                    'raw_address': address,  # Endereço original para referência
                    'lat': lat,
                    'lon': lon,
                    'stop': stop,
                    'bairro': bairro,
                    'city': city,
                    'customer': customer,
                    'phone': phone,
                    'tracking': tracking
                })
                
                stop_counter += 1
        
        return addresses
        
    except Exception as e:
        raise Exception(f"Erro ao parsear Excel Shopee: {str(e)}")


class ShopeeRomaneioParser:
    """Parser compatível com código legado"""
    
    @staticmethod
    def parse(file_path: str) -> List[ShopeeDelivery]:
        """Parse e retorna lista de ShopeeDelivery objects"""
        data = parse_shopee_excel(file_path)
        
        deliveries = []
        for item in data:
            delivery = ShopeeDelivery(
                tracking=item['tracking'],
                address=item['address'],
                bairro=item.get('bairro', ''),
                city=item.get('city', ''),
                latitude=item.get('lat', 0.0),
                longitude=item.get('lon', 0.0),
                stop=item.get('stop', 0),
                customer_name=item.get('customer', ''),
                phone=item.get('phone', '')
            )
            deliveries.append(delivery)
        
        return deliveries
